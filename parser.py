from pdf2image import convert_from_path
from PIL import Image
import pytesseract
from PyPDF2 import PdfWriter, PdfReader
import io
import pdfplumber
import re
from datetime import datetime
import re
import copy
import PyPDF2
import pandas as pd
import csv

import os
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

import openpyxl
from openpyxl import Workbook
import os
from urllib.parse import quote

class Parser:
    def __init__(self,file,link):
        self.table_data = []
        
        self.file = file
        
        self.total_text = ''
        self.lines = []
        
        self.pages = 0

        self.claims_table = [] # instantiate with at least 15 (if greater than 15 add to excel file and empty again)

        self.consistent_info = {}
        self.claims_data = {'etf_number': '',
              'claim_id': '',
              'patient_name': '',
              'patient_id': '',
              'provider_name': '',
              'provider_id': '',
              'date_of_service': '',
              'payer_name': '',
              'payer_id': '',
              'billed_amount': '',
              'amount_allowed': '',
              'paid_amount': '',
              'patient_responsibility': '',
              'denial_code': '', # i think this is for remark/payer code potentially
              'denial_reason_description': '',
              'adjustment_code': '', 
              'adjustment_reason_desc': '',
              'appeal_status': '',
              'appeal_date': '',
              'appeal_decision_date': '',
              'appeal_outcome': '',
              'remittance_date': '',
              'service_line_details': ''}

        self.service_lines = {'service_line_id': [],
                              'claim_id': [],
                              'service_code': [],
                              'units': [],
                              'line_item_billed_amount': [],
                              'line_item_allowed_amount': [],
                              'line_item_paid_amount': [],
                              'cagc': [],
                              'carc': [],
                              'remark_code': []}

        self.codes = {'REMARK': {'': ''}, 'GROUP': {'': ''}, 'CLAIM ADJUSTMENT': {'': ''}, 'CLAIM STATUS': {'': ''}}
        self.curr_code = ""
        self.hyperlink = link

    def reset_data(self):
        self.table_data = []
        self.empty_list()
        self.clear_service_lines_table()
        self.total_text = '' 

    def flatten(self, xs): 
        # Initialize list for this layer 
        flat_list = []
        for x in xs: 
            # If it's a list, recurse down and return the interior list
            if isinstance(x, list): 
                flat_list += self.flatten(x)
            # Otherwise, add to this layer's list
            else: 
                flat_list.append(x) 
        return flat_list

    def change_to_code(self):

        for n in self.claims_table:
            for i in n:
                if i == 'denial_reason_description':
                    n[i] = [self.codes['REMARK'][j] for j in n[i]]

                if i == 'service_line_details':
                    if n[i]:
                        n[i]['cagc'] = self.flatten(n[i]['cagc'])
                        n[i]['carc'] = self.flatten(n[i]['carc'])

                        # include the reasoning
                        n[i]['carc_description'] = [self.codes['CLAIM ADJUSTMENT'][j] for j in n[i]['carc']]
                        n[i]['cagc_description'] = [self.codes['GROUP'][j] for j in n[i]['cagc']]

                if i == 'adjustment_reason_desc':
                    n[i] = [self.codes['REMARK'][j] for j in n['adjustment_code']]


    def find_last_row(self, filename, sheet_name):
        # Load the existing workbook
        workbook = load_workbook(filename)
        
        # Select the active sheet (or a specific sheet)
        sheet = workbook[sheet_name]
        
        # Find the last row
        return sheet.max_row
    def convert_to_excel(self):
        self.change_to_code()
        # we want to change everything to the codes first
        df = pd.DataFrame(self.claims_table)

        #df.to_excel('claims_table.xlsx', index=False)

        filename = 'claims_table.xlsx'

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                df.to_excel(writer, index = False)
                new_sheet_name = list(writer.sheets.keys())[-1]
        else:
            output = df
            output.to_excel(filename, index = False)
            new_sheet_name = 'Sheet1'

        workbook = load_workbook(filename)
        sheet = workbook[new_sheet_name]

        for row in range(2, sheet.max_row + 1):
            etf_number_cell = sheet.cell(row=row, column=1)  # First column of each row
            etf_number_value = etf_number_cell.value

            # Add hyperlink only if the cell doesn't already have one
            if etf_number_value is not None and not etf_number_cell.hyperlink:
                etf_number_cell.hyperlink = ''
                etf_number_cell.style = "Hyperlink"

        workbook.save(filename)

    def convert_quicker(self):
        pages_iterated = 1
        # i think we should iterate over every 5 pages and then test
        with pdfplumber.open(self.file, laparams = {"detect_vertical": False}) as pdf:
            # iterate over each page
            for index in range(len(pdf.pages)):
                page = pdf.pages[index]

                # get relevant info from first page
                if index == 1: self.extract_first_page()

                # extract text
                text = page.extract_text()
                if text: self.total_text += text
                self.extract_table_data(page)

        if self.consistent_info == {}:
            self.extract_first_page()

        # place into new function
        self.parse_through_text()
        self.pages = 1
        self.reset_data()
        self.convert_to_excel()

    def extract_table_data(self, page):
        extracted = []

        extracted = page.extract_tables()
        for table in range(len(extracted)):
            for row in range(len(extracted[table])):
                if row is not None:
                    # Replace newline characters with an empty string in each element of the row
                    extracted[table][row] = [cell.replace('\n', ' ') if cell is not None else '' for cell in extracted[table][row]]
                else:
                    print("Empty row encountered")

            self.table_data.append(extracted[table])

    def print_text(self):
        print(self.total_text)

    def empty_list(self):
        self.claims_data = {n: '' for n in self.claims_data}

    def clear_service_lines_table(self):
        self.service_lines = {n: [] for n in self.service_lines}

    def split_res(self,delimiters,line):
            regex_pattern = '|'.join(map(re.escape, delimiters))
            res = re.split(regex_pattern, line)
            return res

    def extract_first_page(self):
        if self.table_data:
            self.consistent_info['payer_name'] = self.table_data[0][0][0].split(' - ')[0]

        #print("HERE FIRST PAGE")
        #'payee_tax_id': r'Payee Tax ID:\s*(\d+)',
        # 'payment_amount': r'Payment Amount:\s*([\d.]+)',
        regex_patterns = {
            'payer_id': r'Payee ID:\s*(\d+)',
            'etf_number': r'Check/EFT Trace Number:\s*(\d+)',
            'remittance_date': r'Check/EFT Date:\s*([\d/]+)',
            'payee_name': r'Payee Name:\s*([\w\s]+)'
        }
        # Initialize a dictionary to store the extracted information
        extracted_info = {}
        #print(self.table_data)

        # Iterate over the patterns and apply them to the second element of the list
        for key, pattern in regex_patterns.items():
            match = re.search(pattern, self.table_data[0][0][1])
            if match:
                self.consistent_info[key] = match.group(1)
            
        if "Provider Adjustment Code" == self.table_data[1][0][0]:
            self.consistent_info['adjustment_code'] = [self.table_data[1][n][0] for n in range(1, len(self.table_data[1]))]
            self.consistent_info['adjustment_reason_desc'] = [self.table_data[1][n][1] for n in range(1, len(self.table_data[1]))]
        else:
            #print("here")
            self.consistent_info['adjustment_code'] = ''
            self.consistent_info['adjustment_reason_desc'] = ''

    def fill_in(self):
        # change after to use a for loop
        for n in self.consistent_info:
            self.claims_data[n] = self.consistent_info[n]
    
    def check_code_description(self, n):
        # we deciding on the if it is a CLAIM ADJUSTMENT REASON CODE(S): or a CLAIM STATUS CODE(S):
        # or REMARK CODE(S):
        curr = {'CLAIM ADJUSTMENT REASON CODE(S):': 'CLAIM ADJUSTMENT', 'CLAIM STATUS CODE(S):': 'CLAIM STATUS', 'REMARK CODE(S):': 'REMARK', 'GROUP CODE(S)': 'GROUP'}

        for i in curr:
            if i in n:
                self.curr_code = curr[i]

        # otherwise it is a number
        # we wanna split the text by =
        code_des = n.split('=')
        if len(code_des) == 2:
            self.codes[self.curr_code][code_des[0]] = code_des[1]

    def parse_through_text(self):

        # need to account for other pages than just the first one
        if self.consistent_info['adjustment_code'] != '' and self.pages == 0:
            list_instance = 2
        else:
            list_instance = 1
        
        get_lines = self.total_text.split('\n')
        if_code = 0

        for n in get_lines:
            if if_code: self.check_code_description(n)

            if "Patient Name" in n:
                #self.add_to_table()
                self.fill_in()
                res = self.split_res(["Patient Name:", "Claim Number:", "Claim Date:", "Claim Status Code:"], n)
                self.claims_data['patient_name'] = res[1]
                self.claims_data['claim_id'] = res[2]
                self.claims_data['date_of_service'] = res[3]

            if 'Patient ID' in n:
                res = self.split_res(["Patient ID:", "Group / Policy:", "Facility Type:", "Claim Charge:"], n)
                self.claims_data['patient_id'] = res[1]
                self.claims_data['billed_amount'] = res[4]

            if 'Patient Ctrl Nmbr' in n:
                res = self.split_res(["Patient Ctrl Nmbr:", "Contract Hdr:", "Claim Frequency:", "Claim Payment:"], n)
                self.claims_data['amount_allowed'] = res[4]
                self.claims_data['patient_responsibility'] = res[4]

            if 'Rendering Prvd' in n:
                res = self.split_res(["Rendering Prvd:", "Rendering Prv ID:", "Claim Received Date:", "Patient Resp:"], n)
                self.claims_data['provider_name'] = res[1]
                self.claims_data['provider_id'] = res[2]
                self.claims_data['paid_amount'] = res[4]

            if 'Original Ref Nmbr:' in n:
                res = self.split_res(["Original Ref Nmbr:"], n)

            if 'Code Descriptions' in n: if_code = 1

            if 'Line Details' in n:
                # ok my thought rn is to check if claims_data == "" then we want to append to previous list
                # i need to account for the other table
                
                if self.table_data[list_instance][0][0] == 'Payer: AETNA BETTER HEALTH OF ILLINOIS - MEDICAID':
                    list_instance+=1
                # need to do line details
                
                table = self.table_data[list_instance]
                if len(table) > 1:
                    for n in range(1,len(table)):
                        #print(table[n])
                        self.service_lines['service_line_id'].append(table[n][0])
                        self.service_lines['claim_id'].append(self.claims_data['claim_id'])
                
                        table[n][5] = table[n][5].split('/')
                        self.service_lines['service_code'].append(table[n][5][0]if len(table[n][5]) == 3 else '')
                        self.service_lines['units'].append(table[n][5][2]if len(table[n][5]) == 3 else '')
                        
                        self.service_lines['line_item_billed_amount'].append(table[n][8])
                        # difference betwen billed and allowed amount  
                        # try something difference cuz this is not working      
                        self.service_lines['line_item_paid_amount'].append(table[n][11])

                        table[n][9] = table[n][9].split('-')
                        
                        if len(table[n][9]) > 1:
                            table[n][9][1] = table[n][9][1].split(' ')
                            self.service_lines['cagc'].append(table[n][9][0])
                            self.service_lines['carc'].append(table[n][9][1][0])
                        else:
                            self.service_lines['cagc'].append('')
                            self.service_lines['carc'].append('')


                        table[n][6] = '' if table[n][6] == '' else table[n][6].split(' ')
                        self.service_lines['remark_code'].append(table[n][6])
                        # add to denial code reasoning

                        # add adjustment reasons after
                        #self.claims_data['']

                if self.claims_data['patient_name'] == '':
                    # Ensure that 'denial_code' and 'denial_reason_description' fields are initialized properly
                    if 'denial_code' not in self.claims_table[-1]:
                        self.claims_table[-1]['denial_code'] = []
                    if 'denial_reason_description' not in self.claims_table[-1]:
                        self.claims_table[-1]['denial_reason_description'] = set()  # assuming it is a set

                    # Extend the existing service line details
                    for i in self.service_lines:
                        self.claims_table[-1]['service_line_details'][i].extend(self.service_lines[i].copy())

                    # Collect unique denial codes
                    reasons = set()
                    for n in self.service_lines['remark_code']:
                        for i in n:
                            reasons.add(i)

                    # Update 'denial_code' and 'denial_reason_description'
                    self.claims_table[-1]['denial_code'].update(reasons.copy())  # Extend list with reasons
                    self.claims_table[-1]['denial_reason_description'].update(reasons.copy())  # Use update() instead of union()

                    # Print the final data to confirm changes
                    #print(self.claims_table[-1]['denial_code'])
                    #print(self.claims_table[-1]['denial_reason_description'])
                    #print(self.claims_table[-1])
                else:
                    
                    # for this i am only gonna collect the set
                    reasons = set()

                    for n in self.service_lines['remark_code']:
                        for i in n:
                            reasons.add(i)
                    self.claims_data['denial_code'] = reasons
                    self.claims_data['denial_reason_description'] = reasons

                    self.claims_data['service_line_details'] = self.service_lines.copy()
                    self.add_to_table()

                list_instance+=1

        #self.add_to_table()
                
    def add_to_table(self):
        self.claims_table.append(self.claims_data.copy())
        self.clear_service_lines_table()
        self.empty_list()          

#-----------------------------------------------------------------------------------------------------------
#-----------------------------------------------------------------------------------------------------------
# now the goal is to test bigger
# can place into an excel file

# adjustments
# figure out how to place 2 pdfs into the same instance of the class
# then figure out how to merge using pythonk

# we need a list of hyperlinks to allow for the hyperlink of all of themsssssssssss

def merge_files():
    # Load the original Excel file
    excel_file = 'claims_table.xlsx'
    wb = openpyxl.load_workbook(excel_file, data_only=False)

    # Create a new workbook to store combined data
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Combined"

    # Initialize a variable to keep track of the current row in the combined sheet
    current_row = 1

    # Iterate over all sheets and append their contents to the new worksheet
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Optionally, write the sheet name as a header in the combined worksheet
        #new_ws.cell(row=current_row, column=1, value=f"Sheet: {sheet}")
        #urrent_row += 1

        # Iterate over all rows and columns in the current sheet
        for row in ws.iter_rows(values_only=False):
            new_row = []

            for cell in row:
                # Copy the cell value to the new combined sheet
                new_cell = new_ws.cell(row=current_row, column=cell.col_idx, value=cell.value)

                # Copy hyperlinks, if present
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink.target

                # Copy the cell's style
                new_cell.style = cell.style

                # Copy formulas, if present
                if cell.data_type == "f":
                    new_cell.value = f"={cell.value}"

                # Copy set or list-like data, if any
                if isinstance(cell.value, (list, set)):
                    new_cell.value = str(cell.value)

            # Move to the next row in the combined sheet
            current_row += 1

        # Optionally, leave a blank row between sheets' contents
        #current_row += 1

    # Save the new workbook with combined sheets and hyperlinks
    new_wb.save('combined_sheets_with_hyperlinks.xlsx')


folder_path = "read_files"

for pdf_name in os.listdir(folder_path):
    if pdf_name.endswith('.pdf'):  # Process only PDF files
        pdf_path = os.path.join(folder_path, pdf_name)
        #abs_path = os.path.abspath(pdf_path)
        file_url = ""
        #print(file_url)
        parsed_file = Parser(pdf_path, file_url)
        parsed_file.convert_quicker()

'''
links = ['']

for n in range(1):
    print(n)
    pdf_name = str(n)+'.pdf'
    link = links[n]
    parsed_file = Parser(pdf_name, link)
    parsed_file.convert_quicker()

'''
merge_files()




